# Create insert statements for RORALGS
# reads excel with roralgs values at <FILE HERE> and creates an insert statement
# written by Justin DeKock 11/7/2025 

import sys
from openpyxl import load_workbook, workbook, worksheet
from datetime import datetime as dt

FTIME = dt.strftime(dt.now(), "%m%d%Y_%H%M%S")
T_DIR = rf"T:\Enrollment Retention Management\Student_Financial_Services\SYSTEMS\buddy_pb_sql_fall25\roralgs"
EXL_T = rf"{T_DIR}\roralgs_draft.xlsx"
SQL_T = rf"{T_DIR}\sql\roralgs_ins_{FTIME}.sql"
TDRIVE = {"exl": EXL_T, "sql": SQL_T}

EXL_L = rf"ref\roralgs_dev.xlsx"
SQL_L = rf"sql\roralgs_ins_{FTIME}.sql"
LOCAL = {"exl": EXL_L, "sql": SQL_L}

# open excel file
def open_xl(f_name: str) -> workbook:
    return load_workbook(f_name, read_only=True, data_only=True)

# returns rows of a worksheet as a tuple after opening workbook
def read_sheet(ws: worksheet) -> tuple:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
    return tuple(rows)

# returns each row as a dict, with row number mapped to a dict with hdr mapped to the data
def parse_rows(rows: tuple) -> dict:
    row_data = dict()
    for ir, row in enumerate(rows[1:]):
        data = dict()
        for ic, hdr in enumerate(rows[0]):
            data[hdr] = row[ic]
        row_data[ir + 1] = data
    return row_data
    
# accept dict created in parse_rows, build insert statement string        
def build_insert_stmnt(data: dict) -> str:
    # get tuples of row strings - each contains a string to concat 
    # to insert statement, represents one row of data being inserted
    rows = rows_to_insert(data)
    # keys are the column headers from excel sheet - AIDY_CODE, KEY_1..., AMOUNT
    # COLUMNS MUST BE NAMED THIS WAY, THEY'RE USED TO BUILD RORALGS_KEY_X STYLE
    # FIELD NAMES DYNAMICALLY
    hdrs = data[1].keys()
    # BUILD THE INSERT STATEMENT, .join() UNWRAPS STRING IN TUPLES WITH ', ' AS DELIMINITER
    return (
f"""INSERT INTO FAISMGR.RORALGS 
({', '.join([f'RORALGS_{hdr}' for hdr in hdrs])}) 
VALUES 
{',\n'.join(rows)};""")

# loop through data dictionary (spreadsheet data) and return a tuple of row strings
# to be concatenated to insert statement string
def rows_to_insert(data: dict) -> tuple:
    rows = []
    for d in data.values():
        vals = []
        for k, v in d.items():
            if v is None:
                vals.append("NULL")
                
            elif str(v).isnumeric() and not 'AIDY' in str(k):
                vals.append(str(v))
            else:
                vals.append(f"'{d[k]}'")
        rowstr = f"({', '.join(vals)})"
        rows.append(rowstr)
    return tuple(rows)

# entrypoint
def main():
    env = LOCAL
    if len(sys.argv) > 1:
        match sys.argv[1]:
            case "t":
                env = TDRIVE
            
    wb = open_xl(env["exl"])
    ws = wb["Sheet1"]
    rows = read_sheet(ws)
    data = parse_rows(rows)
    ins = build_insert_stmnt(data)
    print(ins)
    with open(env["sql"], "w") as f:
        f.write(ins)
        
if __name__ == "__main__":
    main()