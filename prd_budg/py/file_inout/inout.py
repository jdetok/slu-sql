from openpyxl import load_workbook, workbook, worksheet
import os

def make_output_dir(dirname):
    try:
        os.mkdir(dirname)
        print(f"Directory '{dirname}' created successfully.")
    except FileExistsError:
        print(f"Directory '{dirname}' already exists.")
    except PermissionError:
        print(f"Permission denied: Unable to create '{dirname}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

def write_to_file(fname: str, content: str):
    with open(fname, "a") as f:
        f.write(content)
        
# open excel file
def open_xl(f_name: str) -> workbook:
    return load_workbook(f_name, read_only=True, data_only=True)

# get tuple of worksheet rows as a tuple after opening workbook
def read_sheet(ws: worksheet) -> dict[str, tuple[tuple]]:
    data_rows = []
    content = dict()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_tup = tuple(row)
        if i == 0: 
            content["cols"] = row_tup
            continue
        data_rows.append(row_tup)
    content["rows"] = tuple(data_rows)
    return content


# join a tuple of vals to a comma separated string, with newlines at limit
# the newlines will start at tab
def fmt_join(arr: tuple, delim=", ", limit=80, tab=0, safe=True) -> str:
    #  read text char by char, track lines, if char idx gets > limit, newline
    txt = ""
    char_idx = tab
    t = "" if tab == 0 else (" " * tab)

    for i, s in enumerate(arr): 
        if s is None:
            continue
        slen = len(s)
        dlim = len(delim)
        if (slen + tab) > limit:
            if safe: 
                raise ValueError(f"lengeth of str + tab ({slen + tab}) must be less than {limit}")
            else: limit = (slen + tab + 1)
                
        
        if (char_idx + slen) >= limit:
            txt += f"\n{t}{s}"
            char_idx = (tab + slen)
        else: 
            txt += s
            char_idx += slen
        
        if i+1 < len(arr):
            txt += delim
            char_idx += dlim
    return txt
            
            
            

