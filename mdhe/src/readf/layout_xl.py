from openpyxl import load_workbook, worksheet
import logging
logger = logging.getLogger(__name__)

# open the file - headers are on row 3 | only need cols 0-5
# returns worksheet object after opening workbook
# if csv_file is passed, it'll open the excel file and save a csv
def open_layout_worksheet(f_name, sh_name) -> worksheet:
    try:
        wb = load_workbook(f_name, read_only=True, data_only=True)
        logger.info(f"successfully loaded {f_name} | getting data from {sh_name}")
        return wb[sh_name]
    except:
        logger.exception(
            f"fatal error occured attempting to open worksheet "
            f"{sh_name} in workbook at {f_name}"
        )
        raise
    
# return rows of data in the worksheet. data_row should be the first row under
# the headers in the xlsx file
def get_layout_rows(ws, start_row, start_col, end_col) -> tuple:
    try:
        rows = []
        for row in ws.iter_rows(values_only=True, min_row=start_row):
            rows.append(row[start_col:end_col])
        return tuple(rows)
    except: 
        print("error occured converting worksheet data rows to tuple")
        raise